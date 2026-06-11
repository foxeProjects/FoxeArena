import argparse
import csv
from pathlib import Path

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parent.parent
DEFAULT_SOURCE_DIR = Path(r"C:\Users\foxe\Downloads\porra")
CONTACTS_PATH = ROOT / "assets" / "wc-participantes-contactos.csv"
FIELDNAMES = ["nombre", "email", "archivo"]


def clean(value):
    if value is None:
        return ""
    return str(value).strip()


def discover_excel_files(source_dir):
    files = []
    for pattern in ("*.xlsx", "*.xlsm"):
        files.extend(source_dir.glob(pattern))
    return sorted(path for path in files if not path.name.startswith("~$"))


def read_contact(path):
    workbook = load_workbook(path, data_only=True, read_only=True)
    if "HOME" not in workbook.sheetnames:
        raise ValueError("No existe la pestaña HOME")
    sheet = workbook["HOME"]
    return {
      #  "nombre": clean(sheet["C4"].value),
        "email": clean(sheet["C5"].value),
      #  "archivo": path.name,
    }


def write_rows(path, rows):
    with path.open("w", newline="", encoding="utf-8") as file:
        writer = csv.DictWriter(file, fieldnames=FIELDNAMES)
        writer.writeheader()
        writer.writerows(rows)


def main():
    parser = argparse.ArgumentParser(description="Extrae nombre/email desde la pestaña HOME de los Excels de porra.")
    parser.add_argument("--source-dir", default=str(DEFAULT_SOURCE_DIR), help="Carpeta con ficheros .xlsx/.xlsm de porras.")
    parser.add_argument("--output", default=str(CONTACTS_PATH), help="CSV destino de contactos.")
    args = parser.parse_args()

    source_dir = Path(args.source_dir)
    output_path = Path(args.output)
    if not source_dir.exists():
        raise FileNotFoundError(f"No existe la carpeta origen: {source_dir}")

    rows = []
    errors = []
    for excel_file in discover_excel_files(source_dir):
        try:
            rows.append(read_contact(excel_file))
            print(f"OK  {excel_file.name}")
        except Exception as exc:
            errors.append((excel_file.name, str(exc)))
            print(f"ERR {excel_file.name}: {exc}")

    if not rows:
        raise RuntimeError("No se extrajo ningún contacto.")

    write_rows(output_path, rows)
    print(f"CSV actualizado: {output_path}")
    print(f"Contactos importados: {len(rows)}")
    if errors:
        print(f"Ficheros con error: {len(errors)}")


if __name__ == "__main__":
    main()
