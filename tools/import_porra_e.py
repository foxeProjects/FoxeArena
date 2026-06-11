import argparse
import csv
from pathlib import Path

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parent.parent
DEFAULT_SOURCE_DIR = Path(r"C:\Users\Foxe\Downloads\porra")
PARTICIPANTS_PATH = ROOT / "assets" / "wc-participantes-template.csv"
FIELDNAMES = ["participante", "match_num", "group", "date", "team1", "team2", "pred1", "pred2"]

GROUP_MATCH_RANGE = range(1, 73)
GROUP_INPUT_ROWS = range(18, 90)
 
KNOCKOUT_CELL_MAP = {
    **{match_num: f"I{row}" for match_num, row in zip(range(73, 105), range(6, 38))},
    **{match_num: f"I{row}" for match_num, row in zip(range(105, 121), range(41, 57))},
    **{match_num: f"I{row}" for match_num, row in zip(range(121, 129), range(60, 68))},
    **{match_num: f"I{row}" for match_num, row in zip(range(129, 133), range(71, 75))},
    **{match_num: f"I{row}" for match_num, row in zip(range(133, 135), range(78, 80))},
    135: "I85",
    136: "I83",
    137: "I83",
    138: "I84",
    139: "I85",
    140: "C12",
    141: "C10",
    142: "C9",
    143: "C11",
    144: "C13",
}


def clean(value):
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def load_template_rows(path):
    with path.open(newline="", encoding="utf-8") as file:
        rows = list(csv.DictReader(file))
    template = {}
    for row in rows:
        match_num = int(row["match_num"])
        if match_num not in template:
            template[match_num] = {field: clean(row.get(field, "")) for field in FIELDNAMES}
    missing = [match_num for match_num in range(1, 145) if match_num not in template]
    if missing:
        raise ValueError(f"Faltan match_num en la plantilla base: {missing}")
    return template


def build_base_row(template, participant, match_num):
    source = template[match_num]
    return {
        "participante": participant,
        "match_num": str(match_num),
        "group": source["group"],
        "date": source["date"],
        "team1": source["team1"],
        "team2": source["team2"],
        "pred1": "",
        "pred2": "",
    }


def read_home_sheet(path):
    workbook = load_workbook(path, data_only=True, read_only=True)
    if "HOME" not in workbook.sheetnames:
        raise ValueError("No existe la pestaña HOME")
    return workbook["HOME"]


def rows_from_workbook(path, template):
    sheet = read_home_sheet(path)
    participant = clean(sheet["C4"].value)
    if not participant:
        participant = path.stem.strip()

    rows = []
    for match_num, excel_row in zip(GROUP_MATCH_RANGE, GROUP_INPUT_ROWS):
        row = build_base_row(template, participant, match_num)
        row["team1"] = clean(sheet[f"C{excel_row}"].value) or row["team1"]
        row["team2"] = clean(sheet[f"D{excel_row}"].value) or row["team2"]
        row["pred1"] = clean(sheet[f"E{excel_row}"].value)
        row["pred2"] = clean(sheet[f"F{excel_row}"].value)
        rows.append(row)

    for match_num in range(73, 145):
        row = build_base_row(template, participant, match_num)
        row["pred1"] = clean(sheet[KNOCKOUT_CELL_MAP[match_num]].value)
        rows.append(row)

    return rows


def discover_excel_files(source_dir):
    patterns = ("*.xlsx", "*.xlsm")
    files = []
    for pattern in patterns:
        files.extend(source_dir.glob(pattern))
    return sorted(path for path in files if not path.name.startswith("~$"))


def read_existing_rows(path):
    if not path.exists():
        return []
    with path.open(newline="", encoding="utf-8") as file:
        return list(csv.DictReader(file))


def write_rows(path, rows):
    with path.open("w", newline="", encoding="utf-8") as file:
        writer = csv.DictWriter(file, fieldnames=FIELDNAMES)
        writer.writeheader()
        writer.writerows(rows)


def main():
    parser = argparse.ArgumentParser(description="Importa porras desde Excels con pestaña HOME.")
    parser.add_argument("--source-dir", default=str(DEFAULT_SOURCE_DIR), help="Carpeta con ficheros .xlsx/.xlsm de porras.")
    parser.add_argument("--output", default=str(PARTICIPANTS_PATH), help="CSV destino de participantes.")
    parser.add_argument("--mode", choices=["replace", "append"], default="replace", help="replace regenera el CSV; append añade al final.")
    args = parser.parse_args()

    source_dir = Path(args.source_dir)
    output_path = Path(args.output)
    if not source_dir.exists():
        raise FileNotFoundError(f"No existe la carpeta origen: {source_dir}")

    template = load_template_rows(output_path)
    excel_files = discover_excel_files(source_dir)
    if not excel_files:
        raise FileNotFoundError(f"No hay ficheros .xlsx/.xlsm en: {source_dir}")

    imported_rows = []
    errors = []
    for excel_file in excel_files:
        try:
            imported_rows.extend(rows_from_workbook(excel_file, template))
            print(f"OK  {excel_file.name}")
        except Exception as exc:
            errors.append((excel_file.name, str(exc)))
            print(f"ERR {excel_file.name}: {exc}")

    if not imported_rows:
        raise RuntimeError("No se importó ningún participante.")

    rows_to_write = imported_rows
    if args.mode == "append":
        rows_to_write = read_existing_rows(output_path) + imported_rows

    write_rows(output_path, rows_to_write)
    print(f"CSV actualizado: {output_path}")
    print(f"Filas importadas: {len(imported_rows)}")
    print(f"Participantes importados: {len(imported_rows) // 144}")
    if errors:
        print(f"Ficheros con error: {len(errors)}")


if __name__ == "__main__":
    main()
